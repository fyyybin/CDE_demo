# -*- coding: UTF-8 -*-
"""
时间2020-12-14 19：00
"""
import pandas as pd
import os
import json
from openpyxl import Workbook, load_workbook

def No_fill(t):
    if t == None:
        return '暂无'
    else:
        return t
def main():
    dict = []
    filePath = "D:\\Git\\CDE_demo\\data\\"
    file_list = os.listdir(filePath)
    print(file_list)
    for file in file_list:
        path = "D:\\Git\\CDE_demo\\data\\" + file
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.max_row#总行数
        columns = ws.max_column#总列数
        print(file)


        documents = {}
        documents['登记号'] = ws['B1'].value.replace(' ','')
        documents['药物名称'] = ws['B7'].value.replace(' ','')
        documents['药物类型'] = No_fill(ws['B8'].value)
        documents['适应症'] = ws['B10'].value.replace(' ','')
        documents['试验专业题目'] = ws['B11'].value.replace(' ','')
        documents['试验通俗题目'] = ws['B12'].value.replace(' ','')

        mess = {}
        for i in range(17, rows):
            if ws.cell(row=i, column=1).value == '联系人姓名':
                documents['联系人姓名'] = No_fill(ws.cell(row=i, column=2).value)
            if ws.cell(row=i, column=3).value == '联系人座机':
                documents['联系人座机'] = No_fill(ws.cell(row=i, column=4).value)
            if ws.cell(row=i, column=1).value == '联系人手机':
                documents['联系人手机'] = No_fill(ws.cell(row=i, column=2).value)
            if ws.cell(row=i, column=3).value == '联系人Email':
                documents['联系人Email'] = No_fill(ws.cell(row=i, column=4).value)
            if ws.cell(row=i, column=1).value == '联系人邮政地址':
                documents['联系人邮政地址'] = No_fill(ws.cell(row=i, column=2).value)
            if ws.cell(row=i, column=3).value == '联系人邮编':
                documents['联系人邮编'] = No_fill(ws.cell(row=i, column=4).value)

            if ws.cell(row=i, column=1).value == '1、试验目的':
                documents['试验目的'] = ws.cell(row=i+1, column=1).value.replace(' ', '')
            elif ws.cell(row=i, column=1).value == '试验分类':
                mess['试验分类'] = No_fill(ws.cell(row=i, column=2).value)
            elif ws.cell(row=i, column=1).value == '试验分期':
                mess['试验分期'] = No_fill(ws.cell(row=i, column=2).value)
            elif ws.cell(row=i, column=1).value == '设计类型':
                mess['设计类型'] = No_fill(ws.cell(row=i, column=2).value)
            elif ws.cell(row=i, column=1).value == '随机法':
                mess['随机法'] = No_fill(ws.cell(row=i, column=2).value)
            elif ws.cell(row=i, column=1).value == '盲法':
                mess['盲法'] = No_fill(ws.cell(row=i, column=2).value)
            elif ws.cell(row=i, column=1).value == '试验范围':
                mess['试验范围'] = No_fill(ws.cell(row=i, column=2).value)
                key_1 = i #受试者开始的标签
            if ws.cell(row=i, column=1).value == '4、试验分组':
                key_2 = i+1  # 受试者结束的标签
            if ws.cell(row=i, column=1).value == '5、终点指标':
                key_3 = i+1  # 试验分组结束的标签
            if ws.cell(row=i, column=1).value == '6、数据安全监查委员会（DMC）':
                key_4 = i+1 # 终点指标结束标签
            if ws.cell(row=i, column=1).value == '六、试验状态':
                key_5 = i # 试验状态开始标签
            if ws.cell(row=i, column=1).value == '七、临床试验结果摘要':
                key_6 = i+1 # 试验状态结束标签
        documents['试验设计'] = mess

        # 下面是受试者信息字典(key_1, key_2)
        mess_man = {}
        state =  []
        for i in range(key_1, key_2):
            if ws.cell(row=i, column=1).value == '年龄':
                mess_man['年龄'] = ws.cell(row=i+1, column=1).value.replace(' ', '')
            elif ws.cell(row=i, column=1).value == '性别':
                mess_man['性别'] = ws.cell(row=i+1, column=2).value.replace(' ', '')
            elif ws.cell(row=i, column=1).value == '健康受试者':
                mess_man['健康受试者'] = ws.cell(row=i + 1, column=2).value.replace(' ', '')
            if ws.cell(row=i, column=1).value == '入选标准':
                for j in range(i, key_2):
                    state.append(ws.cell(row=j, column=2).value.replace(' ', ''))
                    if ws.cell(row=j+1, column=1).value == '排除标准':
                        mess_man['入选标准'] = state
                        state = []
                        break
            elif ws.cell(row=i, column=1).value == '排除标准':
                for j in range(i, key_2):
                    state.append(ws.cell(row=j, column=2).value.replace('\xa0', ''))
                    if ws.cell(row=j+1, column=1).value == '4、试验分组':
                        mess_man['排除标准'] = state
                        state = []
                        documents['受试者信息'] = mess_man
                        break

        #下面是试验分组(key_2-1， key_3)
        p = 0
        mess = {}
        mess_number, mess_test = {}, {}
        n_1, n_2 = 0, 0
        for i in range(key_2, key_3-1):
            if ws.cell(row=i, column=2).value == '序号':
                if n_1 == n_2:
                    n_1 = i
                else:
                    n_2 = i
            if ws.cell(row=i, column=2).value == None:
                p = 1
        #2020之前的模板
        if p == 0:
            for x in range(n_1, n_2):
                if ws.cell(row=x, column=2).value != '序号':
                    mess_test['名称'] = ws.cell(row=x, column=4).value.replace('中文通用名', '').replace('：','')
                    mess_test['用法'] = ws.cell(row=x, column=5).value.replace('用法用量：', '')
                    mess_number[ws.cell(row=x, column=2).value] = mess_test
                    mess_test = {}
            mess['试验药'] = mess_number
            mess_number = {}
            for x in range(n_2, key_3-1):
                if ws.cell(row=x, column=2).value != '序号':
                    mess_test['名称'] = ws.cell(row=x, column=4).value.replace('中文通用名', '').replace('：', '')
                    mess_test['用法'] = ws.cell(row=x, column=5).value.replace('用法用量：', '')
                    mess_number[ws.cell(row=x, column=2).value] = mess_test
                    mess_test = {}
            mess['对照药'] = mess_number
            mess_number = {}
        #2020之后的复杂模板
        else:
            doc1, doc2 = [], []
            t = 0
            for x in range(n_1+1, n_2):
                if ws.cell(row=x, column=2).value in [2,9]:
                    mess_test['名称'] = doc1
                    mess_test['用法'] = doc2
                    mess_number[ws.cell(row=x, column=2).value - 1] = mess_test
                    doc1, doc2 = [], []
                    mess_test = {}
                    t = ws.cell(row=x, column=2).value - 1
                doc1.append(ws.cell(row=x, column=4).value)
                doc2.append(ws.cell(row=x, column=5).value)
            mess_test['名称'] = doc1
            mess_test['用法'] = doc2
            mess_number[t + 1] = mess_test
            mess['试验药'] = mess_number

            mess_number = {}
            mess_test = {}
            doc1, doc2 = [], []
            for x in range(n_2+1, key_3-1):
                if ws.cell(row=x, column=2).value in [2,9]:
                    mess_test['名称'] = doc1
                    mess_test['用法'] = doc2
                    mess_number[ws.cell(row=x, column=2).value - 1] = mess_test
                    doc1, doc2 = [], []
                    mess_test = {}
                doc1.append(ws.cell(row=x, column=4).value)
                doc2.append(ws.cell(row=x, column=5).value)
            mess_test['名称'] = doc1
            mess_test['用法'] = doc2
            mess_number[t + 1] = mess_test
            mess['对照药'] = mess_number
            mess_test = {}
            mess_number = {}
        documents['试验分组'] = mess

        # 下面是终点指标(key_3-1， key_4)
        mess = {}
        mess_number, mess_test = {}, {}
        n_1, n_2 = 0, 0
        for i in range(key_3 - 1, key_4):
            if ws.cell(row=i, column=2).value == '序号':
                if n_1 == n_2:
                    n_1 = i
                else:
                    n_2 = i

        for x in range(n_1, n_2):
            if ws.cell(row=x, column=2).value != '序号':
                mess_test['指标'] = ws.cell(row=x, column=3).value
                mess_test['评价时间'] = ws.cell(row=x, column=6).value
                mess_test['终点指标选择'] = ws.cell(row=x, column=7).value.replace('\t', '').replace(' ', '')
                mess_number[ws.cell(row=x, column=2).value] = mess_test
        mess['主要终点指标及评价时间'] = mess_number
        mess_number, mess_test = {}, {}
        for x in range(n_2, key_4 - 1):
            if ws.cell(row=x, column=2).value != '序号':
                mess_test['指标'] = ws.cell(row=x, column=3).value
                mess_test['评价时间'] = ws.cell(row=x, column=6).value
                mess_test['终点指标选择'] = ws.cell(row=x, column=7).value.replace('\t', '').replace(' ', '')
                mess_number[ws.cell(row=x, column=2).value] = mess_test
                mess_test = {}
        mess['次要终点指标及评价时间'] = mess_number
        documents['终点指标'] = mess
        documents['数据安全监查委员会（DMC）'] = ws.cell(row=key_4 - 1, column=2).value.replace(' ', '')
        documents['为受试者购买试验伤害保险'] = ws.cell(row=key_4, column=2).value.replace(' ', '')

        # 下面是试验状态(key_5， key_6)
        mess = {}
        for i in range(key_5, key_6):
            if ws.cell(row=i, column=1).value.replace(' ', '') == '1、试验状态':
                mess['试验状态'] = ws.cell(row=i + 1, column=1).value
            elif ws.cell(row=i, column=1).value.replace(' ', '') == '2、试验人数':
                mess_test = {}
                mess_test['目标入组人数'] = No_fill(ws.cell(row=i + 1, column=2).value)
                mess_test['已入组人数'] = ws.cell(row=i + 2, column=2).value
                mess_test['实际入组总人数'] = ws.cell(row=i + 3, column=2).value
                mess['试验人数'] = mess_test

            elif ws.cell(row=i, column=1).value.replace(' ', '') == '3、受试者招募及试验完成日期':
                mess_test = {}
                mess_test['第一例受试者签署知情同意书日期'] = ws.cell(row=i + 1, column=2).value
                mess_test['第一例受试者入组日期'] = ws.cell(row=i + 2, column=2).value
                mess_test['试验完成日期'] = ws.cell(row=i + 3, column=2).value
                mess['受试者招募及试验完成日期'] = mess_test
        documents['试验状态'] = mess
        print('-----------------------------------------------------------')
        dict.append(documents)
    savefile(dict)

def savefile(dic):
    with open('1.json', 'a', encoding='utf-8') as f:
        json_str = json.dumps(dic, indent=4, ensure_ascii=False)
        f.write(json_str)

main()